
import pandas as pd
import yfinance as yf
import openpyxl # Although pandas uses it, importing can be explicit
import time # Potentially for delays if needed

def read_input_data(file_path):
    """
    Reads input data (tickers, dates, constraints) from the specified Excel file
    using openpyxl for precise cell access.

    Args:
        file_path (str): The path to the input Excel file.

    Returns:
        tuple: A tuple containing:
            - list: Tickers.
            - pd.Timestamp: Historical start date.
            - pd.Timestamp: Historical end date.
            - pd.Timestamp: Optimization date.
            - pd.Timestamp: Analysis start date.
            - pd.Timestamp: Analysis end date.
            - pd.DataFrame: Constraints (ticker, min_weight, max_weight).
            Or None for all if an error occurs.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True) # data_only=True reads cell values, not formulas

        # --- Read from Sheet3 ---
        sheet3 = workbook['Sheet3']

        # Read tickers from the first column, starting from row 2
        tickers = []
        # Iterate through rows in column A (index 1) starting from row 2
        for row_index in range(2, sheet3.max_row + 1):
            cell_value = sheet3.cell(row=row_index, column=1).value
            if cell_value: # Only add if cell is not empty
                 tickers.append(str(cell_value)) # Ensure ticker is string
            else:
                 break # Stop if we hit an empty cell in the ticker column

        # Read specific date cells
        start_date = pd.to_datetime(sheet3['B2'].value)
        end_date = pd.to_datetime(sheet3['C2'].value)
        optimization_date = pd.to_datetime(sheet3['E2'].value) # Read the actual date from E1
        analysis_start_date = pd.to_datetime(sheet3['G2'].value)
        analysis_end_date = pd.to_datetime(sheet3['H2'].value)

        # --- Read from Sheet2 ---
        # Use pandas for reading the tabular constraints data, it's more convenient
        constraints_df = pd.read_excel(file_path, sheet_name='Sheet2')
        # Make sure column names match exactly what's in your Excel file
        # Adjust if your columns are named differently
        if not all(col in constraints_df.columns for col in ['Ticker', 'Min Weight', 'Max Weight']):
             print("Error: Sheet2 must contain columns named 'Ticker', 'Min Weight', 'Max Weight'")
             raise ValueError("Missing required constraint columns in Sheet2")

        constraints_df = constraints_df[['Ticker', 'Min Weight', 'Max Weight']]
        constraints_df.columns = ['Ticker', 'min_weight', 'max_weight'] # Standardize column names
        constraints_df['Ticker'] = constraints_df['Ticker'].astype(str) # Ensure tickers are strings for merging later


        print(f"Read data successfully from {file_path}")
        print(f"Tickers: {tickers}")
        print(f"Dates - Hist Start: {start_date}, Hist End: {end_date}, Optim: {optimization_date}, Analysis Start: {analysis_start_date}, Analysis End: {analysis_end_date}")


        return tickers, start_date, end_date, optimization_date, analysis_start_date, analysis_end_date, constraints_df

    except FileNotFoundError:
        print(f"Error: Input file not found at {file_path}")
        return None, None, None, None, None, None, None
    except KeyError as e:
         print(f"Error: Sheet name missing or incorrect: {e}")
         return None, None, None, None, None, None, None
    except Exception as e:
        print(f"Error reading input Excel file: {e}")
        # Print details about the date parsing error if it happens again
        if "datetime" in str(e):
             print("Check that cells B2, C2, E2, G2, H2 in Sheet3 contain valid dates recognized by Excel.")
        return None, None, None, None, None, None, None


def download_price_data(tickers, start_date, end_date):
    """
    Downloads historical adjusted closing prices using yfinance.
    Handles the default auto_adjust=True behavior.

    Args:
        tickers (list): List of stock tickers.
        start_date (pd.Timestamp): Start date for historical data.
        end_date (pd.Timestamp): End date for historical data.

    Returns:
        pd.DataFrame: DataFrame with adjusted closing prices, or None if error.
    """
    print(f"\nDownloading adjusted closing prices for {len(tickers)} tickers from {start_date.date()} to {end_date.date()}...")
    try:
        # Adjust end_date by one day because yfinance excludes the end date
        adj_end_date = end_date + pd.Timedelta(days=1)

        # Use default auto_adjust=True, DO NOT select ['Adj Close'] anymore
        price_data = yf.download(tickers, start=start_date, end=adj_end_date, progress=True)

        # If yf.download returns columns like 'Open', 'High', 'Low', 'Close', etc.
        # because maybe auto_adjust didn't simplify it fully, we select 'Close'.
        # If it already returned a simple DataFrame (Tickers as columns), this won't hurt.
        if isinstance(price_data.columns, pd.MultiIndex):
            price_data = price_data['Close'] # Select the 'Close' column level if multi-level columns exist
        elif 'Close' in price_data.columns and len(price_data.columns) > len(tickers):
             # Handle cases where additional columns like 'Volume' might be present
             # We prioritize the 'Close' column if available alongside others.
             # If only tickers are columns, this condition won't be met.
             if len(tickers) == 1 and 'Close' in price_data.columns:
                 # If only one ticker requested, yfinance might return multiple columns like Open, High, Low, Close etc.
                 price_data = price_data[['Close']].rename(columns={'Close': tickers[0]}) # Select Close and rename column to ticker
             # If multiple tickers are requested and we still see 'Close' it implies MultiIndex which should be handled above
             # This part is tricky, let's stick to the simpler MultiIndex check for now.

        # Handle potential single ticker download (returns Series) after potentially selecting 'Close'
        if isinstance(price_data, pd.Series):
            price_data = price_data.to_frame(name=tickers[0]) # Convert Series to DataFrame

        # Check for failed downloads (all NaN columns)
        failed_tickers = price_data.columns[price_data.isnull().all()].tolist()
        if failed_tickers:
            print(f"Warning: Could not download data for: {failed_tickers}")
            # Drop columns with no data
            price_data = price_data.drop(columns=failed_tickers)

        # Check if any tickers remain
        if price_data.empty:
            print("Error: No price data downloaded successfully for any ticker.")
            return None

        # Check if all requested tickers were successful
        downloaded_tickers = price_data.columns.tolist()
        missing_tickers = [t for t in tickers if t not in downloaded_tickers]
        if missing_tickers:
             print(f"Warning: Failed to get data for: {missing_tickers}")

        print("Price data download complete.")
        # Ensure columns match the valid tickers read initially (order might change)
        valid_tickers = [t for t in tickers if t in downloaded_tickers]
        return price_data[valid_tickers] # Return only valid columns, potentially in original order


    except Exception as e:
        # Add more specific error reporting
        print(f"Error during price data download or processing: {e}")
        import traceback
        traceback.print_exc() # Print detailed traceback for debugging
        return None


def calculate_returns(price_data):
    """
    Calculates daily and weekly returns from price data.

    Args:
        price_data (pd.DataFrame): DataFrame of adjusted closing prices.

    Returns:
        tuple: A tuple containing:
            - pd.DataFrame: Daily returns.
            - pd.DataFrame: Weekly returns (Friday-to-Friday).
    """
    if price_data is None or price_data.empty:
        return None, None

    print("\nCalculating returns...")
    # Daily Returns
    daily_returns = price_data.pct_change().dropna()

    # Weekly Returns (resample to weekly frequency, using Friday)
    # Use .last() to get the last price of the week, then calculate pct change
    weekly_returns = price_data.resample('W-FRI').last().pct_change().dropna()

    print("Returns calculation complete.")
    return daily_returns, weekly_returns



